import json
import sys
import qa_env
import commn
import time
import re
import socket
import jsonHelper
from texttable import Texttable
import multiprocessing
import pafy
import os
import subprocess
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
import settings
import requests
from termios import tcflush, TCIFLUSH

VCO_SCRIPTS_PATH = qa_env._VCO_SCRIPTS_ROOT + str(qa_env._VCO_VERSION)

class Dotable(dict):
    __getattr__= dict.__getitem__
    def __init__(self, d):
        self.update(**dict((k, self.parse(v))
                           for k, v in d.iteritems()))
    @classmethod
    def parse(cls, v):
        if isinstance(v, dict):
            return cls(v)
        elif isinstance(v, list):
            return [cls.parse(i) for i in v]
        else:
            return v

def download(filename):
    url = "https://www.youtube.com/watch?v=lKYVrPhlecI"
    try:
       video = pafy.new(url)
       best = video.getbest()
       filename = best.download(filepath= filename)
    except:
       extype, exval = sys.exc_info()[:2]
       print("critical", "execution failure: " + str(extype) + str(exval))
       return None
    return filename

def reboot(rpc_handle):
    cmd = "reboot"
    ret  = commn.exec_cmd(rpc_handle,cmd,20)
    print ret
    
def get_run_for_case(case_id):
           for case in cases.json():
               if case_id == case['case_id']:
                   return case['id']            #returning run id for given case id

def get_title_for_case(case_id):
          for case in cases.json():
              if case_id == case['case_id']:
                 return case['title']

class USB_Modem(object):
     def __init__(self):
        with open("test_config.json") as json_file:
             self.test_config = Dotable(json.load(json_file)) 
        self.modem_desc = dict()
        self.multi_modem_desc = dict()
        for channel in self.test_config["USBHub"]["channels"]:
            self.modem_desc[channel] = dict()
            self.multi_modem_desc[channel] = dict()
            for port in self.test_config["USBHub"]["ports"]:
                self.modem_desc[channel][port] = dict()

     def start_rpc(self):
         ssh_handle = commn.connect(host=self.test_config["edge"]["hostname"],login='root',password=self.test_config["edge"]["password"],port=22,handle_type="SSH")
         cmd = "nohup /root/qa/check_rpc_server.py --port 7777 &"
         commn.exec_cmd(ssh_handle,cmd,'#',120)

     def Connect_to_Edge(self):
       try:
          self.rpc_handle = commn.connect(self.test_config['edge']['hostname'],port = 7777,mode = None ,ip_port = None)
          print("Info","Connected to Edge")
          print "Getting edge version"
          cmd = "/opt/vc/sbin/edged -v"                     #Edge Image version
          reply = commn.exec_cmd(self.rpc_handle,cmd,60)
          match = re.search(r'Build rev:\s+(\S+)',reply)
          if match:
             self.edge_version = match.group(1)
          return True
       except:
             extype, exval = sys.exc_info()[:2]
             commn.logg_add('ERROR',"critical"+"execution failure: " + str(extype) + str(exval))
             return None
                         
     def EnablePort(self):
         for channel in self.test_config["USBHub"]["channels"]:
             if self.test_config["USBHub"][channel]["status"] == "connected":
                for port in self.test_config["USBHub"]["ports"]:
                    if port == self.active_port:
                       cmd = self.test_config["USBHub"]["name"]+" --action enable"+" --port "+port
                       reply = commn.exec_cmd(self.rpc_handle,cmd,20)
                       print reply
                    else:
                      cmd = self.test_config["USBHub"]["name"]+" --action disable"+" --port "+port
                      reply = commn.exec_cmd(self.rpc_handle,cmd,20)
                      print reply
        
     def EnableChannel(self):
         for channel in self.test_config["USBHub"]["channels"]:
             if channel == self.active_channel:
                cmd = "echo "+self.test_config["edge"][channel]+" > /sys/bus/usb/drivers/usb/bind"
                print cmd
                reply = commn.exec_cmd(self.rpc_handle,cmd,20)
             else:
                cmd = "echo "+self.test_config["edge"][channel]+" > /sys/bus/usb/drivers/usb/unbind" 
                print cmd
                reply = commn.exec_cmd(self.rpc_handle,cmd,20)

     def DisableChannel(self):
         for channel in self.test_config["USBHub"]["channels"]:
                cmd = "echo "+self.test_config["edge"][channel]+" > /sys/bus/usb/drivers/usb/unbind"
                print cmd
                reply = commn.exec_cmd(self.rpc_handle,cmd,20)

     def EnableChannels(self,usb_set): 
         print usb_set
         '''if usb_set == 1:
            cmd = "echo "+self.test_config["edge"]["USB3"]+" > /sys/bus/usb/drivers/usb/unbind"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            print cmd
            print reply
            cmd = "echo "+self.test_config["edge"]["USB4"]+" > /sys/bus/usb/drivers/usb/unbind"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            print cmd
            print reply
            cmd = "echo "+self.test_config["edge"]["USB1"]+" > /sys/bus/usb/drivers/usb/bind"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            print cmd
            print reply
            time.sleep(3)
            cmd = "echo "+self.test_config["edge"]["USB2"]+" > /sys/bus/usb/drivers/usb/bind"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            print cmd
            print reply
         elif usb_set == 2:
            cmd = "echo "+self.test_config["edge"]["USB1"]+" > /sys/bus/usb/drivers/usb/unbind"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            cmd = "echo "+self.test_config["edge"]["USB2"]+" > /sys/bus/usb/drivers/usb/unbind"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            cmd = "echo "+self.test_config["edge"]["USB3"]+" > /sys/bus/usb/drivers/usb/bind" 
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            cmd = "echo "+self.test_config["edge"]["USB4"]+" > /sys/bus/usb/drivers/usb/bind"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)'''

     def check_ip(self,cmd):  
         ip_not_assigned =  1
         ip_count = 0
         while ip_not_assigned:
               time.sleep(5)
               reply = commn.exec_cmd(self.rpc_handle,cmd,20)
               reply = re.search(r'inet addr:(\S+)',reply)
               ip_count += 1
               if reply:
                  ip_not_assigned = 0
               elif ip_count > 7:
                    print "Ip address not assigned for Modem"
                    return None
         try:
            socket.inet_aton(reply.group(1))
            return reply.group(1)
         except socket.error:
            print ("Illegal IP")
            return None

     def modem_insert_remove(self):
        #In a loop activate one port at a time and disable other ports
        for channel in self.test_config["USBHub"]["channels"]:
            if self.test_config["USBHub"][channel]["status"] == "connected":
               for port in self.test_config["USBHub"]["ports"]:
                  if self.test_config["USBHub"][channel][port] == "connected":
                     self.active_channel = channel
                     self.active_port = port
                     print channel
                     print port
                     #self.EnablePort()
                     self.EnableChannel() # enable one channel on the edge
                     print "network restart and sleep"
                     ret = self.basic_test_insert(multi=False) 
                     #reboot
                     self.DisableChannel()  # disable all channels after completing test
                     #print "Rebooting edge to perform reboot test"
                     '''reboot(self.rpc_handle)
                     time.sleep(150)
                     self.start_rpc()
                     status = self.Connect_to_Edge()
                     if status != None:
                        print "Connected to Edge after reboot"
                        ret = self.basic_test_insert()
                     else:
                        print "Could not perform reboot test"'''
                     print "----------TEST COMPLETE----------"
                     #cmd = "vc_procmon restart" 
                     #reply = commn.exec_cmd(self.rpc_handle,cmd,20)
                     print "----Sleeping for 20 seconds before starting next modem ---------"
                     time.sleep(20)
        #Multiple modems test
        self.active_set = "set1"
        self.EnableChannels(usb_set = 1)
        print "sleeping for 5 after activating two channels"
        time.sleep(5)
        ret = self.basic_test_insert(multi=True) 
        return ret
        #disable the active port
        #basic test remove 
     
     def modem_file(self,multi):
          #sleep for few seconds and check modem got detected (/etc/config/modems/modems)
         cmd = "cat /etc/config/modems/modems"
         set_count = 0
         parsed = 0
         multi_parse = 0
         while not parsed:
               length = None 
               while (length <= 200):
                     time.sleep(3)
                     reply = commn.exec_cmd(self.rpc_handle,cmd,20)
                     length = len(reply)
                     print "splitting reply"
               reply = reply.split("\n\n")
               print reply
               for line in reply:
                    print "loop"
                    print line
                    if multi == False:
                       match = re.search('interface\s+(\S+)',line)
                       if match:
                          if match.group(1)[1:-1] != self.active_channel:
                             continue
                          self.modem_desc[self.active_channel][self.active_port]["interface"] = match.group(1)[1:-1]
                          print self.modem_desc[self.active_channel][self.active_port]["interface"]
                       match =  re.search('ifname\s+(\S+)',line)
                       if match:
                          self.modem_desc[self.active_channel][self.active_port]["ifname"] = match.group(1)[1:-1]
                          print self.modem_desc[self.active_channel][self.active_port]["ifname"]
                          parsed = 1       # This parameter is enough to proceed 
                       else:
                          break
                       match = re.search('manufacturer\s+(\S+)',line)
                       if match:
                           self.modem_desc[self.active_channel][self.active_port]["manufacturer"] = match.group(1)
                       match = re.search('product\s+(\S+)',line)
                       if match:
                           self.modem_desc[self.active_channel][self.active_port]["product"] = match.group(1)
                       match = re.search('status\s+(\S+)',line)
                       if match:
                          self.modem_desc[self.active_channel][self.active_port]["status"] = match.group(1)
                       break

                    if multi == True:
                       match =  re.search('interface\s+(\S+)',line)
                       if match:
                          if match.group(1)[1:-1] not in self.test_config["USBHub"][self.active_set]:
                             break   
                          self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["interface"] = match.group(1)[1:-1]
                          #print self.modem_desc[self.active_channel][self.active_port]["interface"]
                       match =  re.search('ifname\s+(\S+)',line)
                       if match:
                          self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["ifname"] = match.group(1)[1:-1]
                          #print self.modem_desc[self.active_channel][self.active_port]["ifname"]
                          multi_parse += 1
                          if multi_parse == 2:
                             parsed = 1
                       match = re.search('manufacturer\s+(\S+)',line)
                       if match:
                           self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["manufacturer"] = match.group(1)
                       match = re.search('product\s+(\S+)',line)
                       if match:
                           self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["product"] = match.group(1)
                       match = re.search('status\s+(\S+)',line)
                       if match:
                          self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["status"] = match.group(1)
                       set_count += 1
         print "modem file"         
         print self.multi_modem_desc      

     def network_file(self,multi):
         cmd = "cat /etc/config/network"
         reply = commn.exec_cmd(self.rpc_handle,cmd,20)
         match = reply.split("\n\n")
         count = 0
         for line in match:
             set_count = 0 
             if multi == False:
                if self.modem_desc[self.active_channel][self.active_port]["ifname"] in line:
                   match = re.search('proto\s+(\S+)',line)
                   if match:
                      self.modem_desc[self.active_channel][self.active_port]["proto"] = match.group(1)        
             if multi == True:
                print "checking" + self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]['ifname']
                if self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]['ifname'] in line:
                   print "found ifnanme in first set"
                   count = set_count
                set_count += 1
                if self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]['ifname'] in line:
                   print "checking" + self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]['ifname']
                   print "found ifname in second set"
                   count = set_count
                print "checked " + self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]['ifname']
                print "count for ifname"
                print count
                match = re.search('proto\s+(\S+)',line)
                if match:
                    self.multi_modem_desc[self.test_config["USBHub"][self.active_set][count]]["proto"] = match.group(1)
                      
         #use interface name for ifconfig <> (Longest time would be 180 seconds) check every 60 secs
         if multi == True:            
             set_count = 0
             cmd  = "ifconfig "+self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]['ifname']
             self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["ip-address"] = self.check_ip(cmd) 
             set_count = 1
             cmd  = "ifconfig "+self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]['ifname']
             self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["ip-address"] = self.check_ip(cmd)
         else:
             cmd = "ifconfig "+self.modem_desc[self.active_channel][self.active_port]["ifname"]
             self.modem_desc[self.active_channel][self.active_port]["ip-address"] = self.check_ip(cmd)
         print self.multi_modem_desc   
         
     def modem_path(self,multi):
         path_not_established = 1
         path_count = 0
         while path_not_established:
               cmd = "debug.py --verbose_path"
               reply = commn.exec_cmd(self.rpc_handle,cmd,20)
               if "Error" in reply or "listening" in reply:
                  print "Server is not listening"
                  continue
               jobj = jsonHelper.jsonToPythonObject(reply)
               for i in xrange(len(jobj)):
                   set_count  = 0
                   if multi == False:
                       if "USB" in  jobj[i]["path"]["interface"]:
                           try: 
                              socket.inet_aton(jobj[i]["path"]["gateway"])
                              self.modem_desc[self.active_channel][self.active_port]["vc-gateway"] = jobj[i]["path"]["gateway"]
                              break
                           except:
                              print "Invalid gateway ip"
                   if multi == True:
                      if self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["interface"] in  jobj[i]["path"]["interface"]:
                         count = set_count
                      set_count += 1
                      if self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["interface"] in  jobj[i]["path"]["interface"]:
                           count = set_count
                      try:
                              socket.inet_aton(jobj[i]["path"]["gateway"])
                              self.multi_modem_desc[self.test_config["USBHub"][self.active_set][count]]["vc-gateway"] = jobj[i]["path"]["gateway"]
                      except:
                              print "Invalid gateway ip"
               if multi == False: 
                  if "vc-gateway" not in self.modem_desc[self.active_channel][self.active_port] and path_count < 10:
                      path_not_established = 1
                      time.sleep(5)
                      path_count += 1
                      if path_count >= 5:
                         cmd  = "usbreset "+self.modem_desc[self.active_channel][self.active_port]["product"]
                         reply = commn.exec_cmd(self.rpc_handle,cmd,20)
                         print reply
                         time.sleep(10)
                      print "path not established vc-gateway-empty"
                      continue
                  else:
                     break
               if multi == True:
                  set_count = 0
                  path_not_established = 0
                  if  "vc-gateway" not in self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]:
                       path_not_established = 1 
                  set_count += 1
                  if "vc-gateway" not in self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]:
                       path_not_established = 1
                  if path_not_established and path_count < 10:
                     time.sleep(5)
                     path_count += 1
                     if path_count >= 5:
                         cmd  = "/etc/init.d/network restart"
                         reply = commn.exec_cmd(self.rpc_handle,cmd,20)
                     print "path not established in multi vc-gateway-empty"
                     continue
                  else:
                    break
         print self.multi_modem_desc         

     def modem_link(self,test,multi):
         link_not_established = 1
         link_count = 0
         while link_not_established:
               cmd = "debug.py --verbose_link"
               reply = commn.exec_cmd(self.rpc_handle,cmd,20)
               print reply
               if "listening" in reply:
                  print "server not listening sleep"
                  time.sleep(10)
                  continue
               jobj = jsonHelper.jsonToPythonObject(reply)
               if "Error" in jobj:
                  print "Error in Link"
                  link_not_established = 1
                  continue
               for i in xrange(len(jobj)):
                   set_count = 0
                   if multi == False:
                      if jobj[i]["type"] == "WIRELESS":
                         if test == 1:
                            return jobj[i]["backupOnly"]
                         if jobj[i]["vpnState"] != "STABLE" and jobj[i]["vpnState"] != "UNSTABLE" and link_count < 10:
                            print "Link unstable ..Waiting"
                            link_count += 1
                            time.sleep(5)
                            link_not_established = 1
                            continue
                         else:
                            link_not_established = 0
                            self.modem_desc[self.active_channel][self.active_port]["name"] = jobj[i]["name"]
                            self.modem_desc[self.active_channel][self.active_port]["state"] = jobj[i]["state"]
                            self.modem_desc[self.active_channel][self.active_port]["vpnState"] = jobj[i]["vpnState"]
                            self.modem_desc[self.active_channel][self.active_port]["logicalId"] = jobj[i]["logicalId"]
                            self.modem_desc[self.active_channel][self.active_port]["publicIpAddress"] = jobj[i]["publicIpAddress"]
                            break
                   if multi == True:
                      set_count = 0
                      if jobj[i]["type"] == "WIRELESS":
                          print jobj[i]["interface"]
                          if self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["interface"] in  jobj[i]["interface"]:
                             count = set_count
                          set_count += 1
                          if self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]["interface"] in  jobj[i]["interface"]:
                              count = set_count 
                          link_not_established = 0
                          set_count  = 0 
                          if "vpnState" not in self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]:
                              link_not_established  = 1
                          set_count += 1
                          if "vpnState" not in self.multi_modem_desc[self.test_config["USBHub"][self.active_set][set_count]]:
                              link_not_established = 1
                          if jobj[i]["vpnState"] != "STABLE" or jobj[i]["vpnState"] != "UNSTABLE" and link_count < 20:
                             print "Link unstable ..Waiting"
                             link_count += 1
                             time.sleep(5)
                             link_not_established = 1
                             continue 
                          else:
                             link_not_established = 0
                             self.multi_modem_desc[self.test_config["USBHub"][self.active_set][count]]["name"] = jobj[i]["name"] 
                             self.multi_modem_desc[self.test_config["USBHub"][self.active_set][count]]["state"] = jobj[i]["state"]
                             self.multi_modem_desc[self.test_config["USBHub"][self.active_set][count]]["vpnState"] = jobj[i]["vpnState"]
                             self.multi_modem_desc[self.test_config["USBHub"][self.active_set][count]]["publicIpAddress"] = jobj[i]["publicIpAddress"]
         print self.multi_modem_desc

     def internet_traffic(self,multi):
         print "checking internet traffic"
         cmd = "ifconfig "+self.test_config["edge"]["ethernet"]+" down"
         reply = commn.exec_cmd(self.rpc_handle,cmd,20)
         cmd = "curl -iv -i https://"+"www.yahoo.com" +" -k"
         reply = commn.exec_cmd(self.rpc_handle,cmd,20)
         match = re.search(r'HTTP/\d.\d\s([0-9]+)',reply)
         if match:
            print match.group()
            if match.group(1) == str(200):
               internet = 'REACHABLE'
            else:
               internet = 'UNREACHABLE'
         else:
              internet  = 'UNREACHABLE'

         if multi == False:
            self.modem_desc[self.active_channel][self.active_port]["internet"] = internet
         elif multi == True: 
            count = 0
            self.multi_modem_desc[self.test_config["USBHub"][self.active_set][count]]["internet"] = internet
            count = 1
            self.multi_modem_desc[self.test_config["USBHub"][self.active_set][count]]["internet"] = internet
         cmd = "ifconfig "+self.test_config["edge"]["ethernet"]+" up"
         reply = commn.exec_cmd(self.rpc_handle,cmd,20)
         print self.multi_modem_desc
         if multi == False:             
            #check dns 
            cmd = "cat /tmp/resolv.conf.auto"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            reply = reply.split("#")
            for i in xrange(len(reply)):
                if self.modem_desc[self.active_channel][self.active_port]["interface"]  in reply[i]:
                      match = re.search(r'nameserver\s([\S]+)',reply[i])
                      if match:
                         print match.group(1)
                         self.modem_desc[self.active_channel][self.active_port]["nameserver"] = match.group(1)

     def route(self,multi):
         cmd = "route -n"
         reply = commn.exec_cmd(self.rpc_handle,cmd,20)
         reply = reply.split("\n")
         for line in reply:
             print self.modem_desc[self.active_channel][self.active_port]["ifname"]
             print line
             if self.modem_desc[self.active_channel][self.active_port]["ifname"] in line:
                match = re.search(r'[\d\.]+\s+([\d.\.]+)',line)
                if match:
                   self.modem_desc[self.active_channel][self.active_port]["gateway"] = match.group(1)
                   break

     def dummy_failover(self,multi):
         filename = "dummy"
         print "Starting dummy failover"
         p = multiprocessing.Process(target = download,args = (filename,))
         p.start()
         if p.is_alive():
            time.sleep(2)
            print "Ethernet link down"
            cmd = "ifconfig "+self.test_config["edge"]["ethernet"]+" down"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            p.join(40)
            p.terminate()
            print "Ethernet link up"
            cmd = "ifconfig "+self.test_config["edge"]["ethernet"]+" up"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)
            cmd  = "ifconfig "+self.test_config["edge"]["ethernet"]
            ip_ethernet = self.check_ip(cmd)
             

     def failover(self,backup,multi):
         filename = "Download "+self.active_channel+self.active_port
         p = multiprocessing.Process(target = download,args = (filename,))
         p.start()
         if p.is_alive():
            time.sleep(2)
            cmd = "ifconfig "+self.test_config["edge"]["ethernet"]+" down"       #Route Metric of ethernet link is always lower
            print "Ethernet link down"
            reply = commn.exec_cmd(self.rpc_handle,cmd,20)                       #compared to Modem
            
            p.join(120)

            if p.is_alive():
                print "Still running ..kill it"
                if backup != 1:
                   self.modem_desc[self.active_channel][self.active_port]["failover"] = "failed"
                else:
                   self.modem_desc[self.active_channel][self.active_port]["backup"] = "failed"
                p.terminate()
                p.join()
            else:
               if os.path.isfile(filename):
                  statinfo = os.stat(filename)
                  if(((statinfo.st_size/1024)/1024) >= 7):
                     if backup != 1:
                        self.modem_desc[self.active_channel][self.active_port]["failover"] = "success" 
                     else:
                        self.modem_desc[self.active_channel][self.active_port]["backup"] = "success"
                  else:
                    if backup != 1:
                       self.modem_desc[self.active_channel][self.active_port]["failover"] = "failed" 
                    else:
                       self.modem_desc[self.active_channel][self.active_port]["backup"] = "failed"
         else:
            print "failover test did not start"
            if backup != 1:
               self.modem_desc[self.active_channel][self.active_port]["failover"] = "failed" 
            else:
               self.modem_desc[self.active_channel][self.active_port]["backup"] = "failed"
         cmd = "ifconfig "+self.test_config["edge"]["ethernet"]+" up"
         reply = commn.exec_cmd(self.rpc_handle,cmd,20)
         cmd = "ifconfig "+self.test_config["edge"]["ethernet"]
         ip_ethernet = self.check_ip(cmd)
         if ip_ethernet == None:
            print "Ethernet link did not come up"
         if os.path.isfile(filename):
            os.remove(filename)       
         print self.modem_desc
         return True  

     def backup_link(self,multi):
         # make usb modem backup only
         backup_wait = 0
         backup = 0
         cmd = VCO_SCRIPTS_PATH+"/"+self.test_config["edge"]["wan_settings"]+" -a modifyinterfaceSetting"+" -v "+self.test_config["edge"]["vco"]+" -enterpriseName "+self.test_config["edge"]["enterprise"]+" -edgeName "+self.test_config["edge"]["name"]+" -configName 'Edge Specific Profile'"+" -configType WAN"+" -wanOverlayName overlay"+" -wanInterface "+self.modem_desc[self.active_channel][self.active_port]["interface"]+" -linkType PUBLIC "+" -wanOverlayType AUTO_DISCOVERED"+" -ethernetType WIRELESS"+" -backUpOnly True"
         os.system(cmd)
         print cmd
         time.sleep(10)

         while(backup == 0): 
            time.sleep(3)
            backup = self.modem_link(1,multi=False)  #check backup link configuration
            backup_wait += 1
            if backup_wait > 5:
               break

         if backup == 1:
            print "Backup link configuration done" 
            self.failover(backup=1,multi=False)   # Do failover test
         else:
            print "Backup link configuration failed"

         cmd = VCO_SCRIPTS_PATH+"/"+self.test_config["edge"]["wan_settings"]+" -a modifyinterfaceSetting"+" -v "+self.test_config["edge"]["vco"]+" -enterpriseName "+self.test_config["edge"]["enterprise"]+" -edgeName "+self.test_config["edge"]["name"]+" -configName 'Edge Specific Profile'"+" -configType WAN"+" -wanOverlayName overlay"+" -wanInterface "+self.modem_desc[self.active_channel][self.active_port]["interface"]+" -linkType PUBLIC "+" -wanOverlayType AUTO_DISCOVERED"+" -ethernetType WIRELESS"+" -backUpOnly False"
         os.system(cmd)
         print cmd
         time.sleep(3)
         backup = self.modem_link(1,multi=False)    #check backup link configuration

     def basic_test_insert(self,multi):
         #select validity of port
         print "Sleeping 20 seconds after channel configuration"
         time.sleep(30)
         self.modem_file(multi)
         #check /etc/config/network interface typ dhcp
         self.network_file(multi) 
         #debug.py --path
         self.modem_path(multi)
         #debug.py --link
         self.modem_link(0,multi)
         #print self.modem_desc 
         #Internet traffic
         self.internet_traffic(multi)
         #check routing table
         self.route(multi)
         #failover test
         self.dummy_failover(multi)
         self.failover(0,multi)
         #backup link
         self.backup_link(multi)
         return True
 
     def basic_test_remove():
         #select validity of port
         # check /etc/config/modems modems
         #check /etc/config/network interface
         # use interface name for ifconfig <> 
         return None     
       
     def reboot():
        #In a loop activate one port
        #basic_test_insert
        #reboot
        #basic_test_insert     
        return None   


testrail = { 22501 : ["internet"],
             22497 : ["product","proto","ip-address"],
             22500 : ["gateway"],
             22502 : ["nameserver"],

             22527 : ["state","vpnState"],
             22530 : "failover",
             22531 : ["backup"]
            }

def get_test_run():
     headers = {"Content-Type": "application/json"}
     cases = requests.get(settings.URL + 'get_tests/' + "726",auth=(settings.USR,settings.PWD),headers=headers)
     print cases.json() 

 
modem = USB_Modem()

p = multiprocessing.Process(target = get_test_run,args = ())
p.start()
p.join(20)
p.terminate()


'''status = modem.start_rpc()
time.sleep(5)
status = modem.Connect_to_Edge()
print status
if status != None:
   status = modem.modem_insert_remove()
   if status == None:
      sys.exit(0)
   else:
      print modem.modem_desc 

   try:
     wb = openpyxl.load_workbook(modem.test_config["edge"]["workbook"])                             
   except:
     print "Workbook does not exist..creating new"
     wb = openpyxl.Workbook()
   
   try: 
      defws = wb.get_sheet_by_name('Sheet')
      wb.remove_sheet(defws)
   except:
      print "Default sheet already removed" 

   sheet_names = wb.get_sheet_names()
   for channel in modem.test_config["USBHub"]["channels"]:
       if modem.test_config["USBHub"][channel]["status"] == "connected":
          for port in modem.test_config["USBHub"]["ports"]: 
             if modem.test_config["USBHub"][channel][port] == "connected":
                prev_title = ((modem.modem_desc[channel][port]["product"]).replace(':','-')).strip('\'')
                if prev_title in sheet_names:
                   print "getting from old sheet"
                   ws = wb.get_sheet_by_name(prev_title)
                   wb.remove_sheet(ws)
               
                print "creating new sheet"
                ws = wb.create_sheet()
                title = str((modem.modem_desc[channel][port]["product"]).replace(':','-'))
                title = title.strip('\'')
                ws.title = str(title)
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 20
                FontStyle = Font(bold=True,size=14) 
                
                row_count = 1
                for key,value in modem.modem_desc[channel][port].items(): 
                     ws.cell(row = row_count,column = 1 ).value = key
                     ws.cell(row = row_count,column = 1 ).font = FontStyle
                     ws.cell(row = row_count,column = 2).value = value 
                     ws.cell(row = row_count,column = 2).font = FontStyle
                     row_count += 1

                result_row_count = 1 
                for key,value in testrail.items():
                     run_id = get_run_for_case(key)                
                     title = get_title_for_case(key)
                     ws.cell(row = result_row_count,column = 4).value = key
                     ws.cell(row = result_row_count,column = 5).value = run_id
                     ws.cell(row = result_row_count,column = 7).value = title
                     print key
                     if key == 22530:
                        print modem.modem_desc[channel][port][value]
                        if "success" in modem.modem_desc[channel][port][value]:
                            print "failover succes"
                            ws.cell(row = result_row_count,column = 6).value = "PASS"
                        else:
                           print "failover failed"
                           ws.cell(row = result_row_count,column = 6).value = "FAIL"
                        result_row_count += 1
                        continue

                     if key == 22527:
                        if "STABLE" in modem.modem_desc[channel][port]["state"] and "STABLE" in modem.modem_desc[channel][port]["vpnState"]:
                            ws.cell(row = result_row_count,column = 6).value = "PASS"
                        else:
                           ws.cell(row = result_row_count,column = 6).value = "FAIL"
                        result_row_count += 1
                        continue

                     if set(value) < set(modem.modem_desc[channel][port].keys()):
                        ws.cell(row = result_row_count,column = 6).value = "PASS"
                     else:
                        ws.cell(row = result_row_count,column = 6).value = "FAIL"
                     result_row_count += 1
        
                             
   wb.save(modem.test_config["edge"]["workbook"])   

   tcflush(sys.stdin, TCIFLUSH)
   user_input = raw_input("Upload results to testrail ? [y/n]: ")
   print user_input
   

   if user_input == 'y':     
      wb = openpyxl.load_workbook(modem.test_config["edge"]["workbook"])
      sheet_names = wb.get_sheet_names()
      print sheet_names 
      
      #Update testrail
      sheet_row_count = 1
      while sheet_row_count <=  result_row_count:
            fail_flag = 0
            success_comment =  " "
            failure_comment = " "
            for sheet in sheet_names:
                ws = wb.get_sheet_by_name(sheet)
                print ws.title
                if ws.cell(row = sheet_row_count,column= 6).value == 'PASS':
                   success_comment  = success_comment +" - " + ws.title +" - "
                   continue
                else:
                   fail_flag = 1
                   print "fail case"
                   failure_comment = failure_comment +" - "+ ws.title+" - "

            if fail_flag == 0:
               print "Updating success result"
               print ws.title
               fields = { "assignedto_id": 2, "comment": success_comment,"status_id": 1,"elapsed": 0,"version": modem.edge_version,"defects": "" }
               print str(ws.cell(row = sheet_row_count,column= 5).value)
               req = requests.post(settings.URL + 'add_result/' + str(ws.cell(row = sheet_row_count,column= 5).value),auth=(settings.USR, settings.PWD), headers=headers,data=json.dumps(fields))
            if fail_flag == 1: 
               print "Update failure result"
               print str(ws.cell(row = sheet_row_count,column= 5).value)
               fields = { "assignedto_id": 2, "comment": failure_comment,"status_id": 5,"elapsed": 0,"version": modem.edge_version,"defects": "" }
               req = requests.post(settings.URL + 'add_result/' + str(ws.cell(row = sheet_row_count,column= 5).value),auth=(settings.USR, settings.PWD), headers=headers,data=json.dumps(fields))
            sheet_row_count += 1
else:
   print "Test Initialisation failed"''' 
